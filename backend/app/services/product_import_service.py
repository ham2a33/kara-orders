from __future__ import annotations

import csv
import io
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.exceptions import ConfigurationError, ValidationAppError
from app.db.models.product import Product
from app.schemas.product import ProductCreateRequest
from app.schemas.product_import import (
    ProductImportConfirmRequest,
    ProductImportConfirmResponse,
    ProductImportError,
    ProductImportParseResponse,
    ProductImportRow,
)
from app.services.ai.openai_provider import OpenAIProvider
from app.services.platform_service import PlatformService
from app.services.product_service import ProductService
from app.services.product_size_parser import normalize_product_name_and_size

_FIELD_ALIASES: dict[str, list[str]] = {
    "name": ["название", "name", "товар", "product", "наименование", "название товара", "наим"],
    "price": ["цена", "price", "стоимость", "цена продажи", "продажа", "цена, тг"],
    "category": ["категория", "category", "группа", "тип"],
    "manufacturer": ["производитель", "manufacturer", "бренд", "brand", "марка"],
    "size": ["размер", "size", "диаметр", "diameter"],
}

_AI_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "price": {"type": "number"},
                    "category": {"type": ["string", "null"]},
                    "manufacturer": {"type": ["string", "null"]},
                    "size": {"type": ["string", "null"]},
                },
                "required": ["name", "price", "category", "manufacturer", "size"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["items"],
    "additionalProperties": False,
}

_AI_INSTRUCTION = (
    "Extract a product catalog from the provided document. "
    "Return each product with name, price, and optional category, manufacturer, and size. "
    "Prices must be numeric. Ignore headers, totals, and non-product rows."
)


class ProductImportService:
    def __init__(self, session: Session, settings: Settings) -> None:
        self.session = session
        self.settings = settings
        self.product_service = ProductService(session=session, settings=settings)

    def parse_excel(self, content: bytes, filename: str) -> ProductImportParseResponse:
        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            sheet_rows = self._read_csv(content)
        elif suffix == ".xlsx":
            sheet_rows = self._read_xlsx(content)
        else:
            raise ValidationAppError("Supported formats: .xlsx, .csv")

        if not sheet_rows:
            raise ValidationAppError("File is empty")

        headers = [str(cell).strip() for cell in sheet_rows[0]]
        if not any(headers):
            raise ValidationAppError("Could not detect column headers")

        mapping = self._auto_map_columns(headers)
        rows: list[ProductImportRow] = []
        for raw_row in sheet_rows[1:]:
            parsed = self._row_from_mapping(headers, raw_row, mapping)
            if parsed is not None:
                rows.append(parsed)

        if not rows:
            raise ValidationAppError("No valid product rows found in file")

        return ProductImportParseResponse(columns=headers, mapping=mapping, rows=rows)

    def parse_pdf(self, content: bytes, filename: str) -> ProductImportParseResponse:
        provider = self._get_ai_provider()
        result = provider.extract_from_file(
            file_bytes=content,
            filename=filename,
            instruction=_AI_INSTRUCTION,
            context="This is a price list or product catalog in PDF format.",
            schema=_AI_SCHEMA,
        )
        return self._parse_ai_result(result.text)

    def parse_photo(self, content: bytes, filename: str, mime_type: str) -> ProductImportParseResponse:
        provider = self._get_ai_provider()
        result = provider.extract_from_image(
            file_bytes=content,
            mime_type=mime_type,
            filename=filename,
            instruction=_AI_INSTRUCTION,
            context="This is a photo of a price list, catalog, or invoice.",
            schema=_AI_SCHEMA,
        )
        return self._parse_ai_result(result.text)

    def confirm_import(self, company_id: UUID, payload: ProductImportConfirmRequest) -> ProductImportConfirmResponse:
        if not payload.rows:
            raise ValidationAppError("No rows to import")

        platform = PlatformService(self.session)
        current_products = self.session.scalar(
            select(func.count(Product.id)).where(Product.company_id == company_id, Product.deleted_at.is_(None))
        ) or 0
        platform.ensure_limit(company_id, "maximum_products", len(payload.rows), current=int(current_products))

        created = 0
        errors: list[ProductImportError] = []

        for index, row in enumerate(payload.rows):
            try:
                self.product_service.create_product(
                    company_id,
                    ProductCreateRequest(
                        name=row.name,
                        size=row.size,
                        price=row.price,
                        category=row.category,
                        manufacturer=row.manufacturer,
                    ),
                )
                created += 1
            except Exception as exc:  # noqa: BLE001 — collect per-row errors for batch import
                errors.append(
                    ProductImportError(
                        row_index=index,
                        name=row.name,
                        message=str(exc),
                    )
                )

        if created:
            platform.log_action(
                action="products_imported",
                company_id=company_id,
                actor_user_id=None,
                resource_type="product",
                resource_id=None,
                description=f"Imported {created} products",
                metadata={"created": created, "errors": len(errors)},
            )

        return ProductImportConfirmResponse(created=created, errors=errors)

    def _get_ai_provider(self) -> OpenAIProvider:
        try:
            return OpenAIProvider(self.settings)
        except ConfigurationError as exc:
            raise ValidationAppError(str(exc)) from exc

    def _parse_ai_result(self, text: str) -> ProductImportParseResponse:
        try:
            payload = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValidationAppError("AI returned invalid data") from exc

        items = payload.get("items")
        if not isinstance(items, list):
            raise ValidationAppError("AI returned no products")

        rows: list[ProductImportRow] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            price = self._parse_price(item.get("price"))
            if not name or price is None:
                continue
            name, size = normalize_product_name_and_size(name, self._optional_str(item.get("size")))
            rows.append(
                ProductImportRow(
                    name=name,
                    price=price,
                    category=self._optional_str(item.get("category")),
                    manufacturer=self._optional_str(item.get("manufacturer")),
                    size=size,
                )
            )

        if not rows:
            raise ValidationAppError("AI could not extract any products")

        return ProductImportParseResponse(rows=rows)

    def _read_csv(self, content: bytes) -> list[list[str]]:
        text = content.decode("utf-8-sig")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        return [[str(cell).strip() for cell in row] for row in reader if any(str(cell).strip() for cell in row)]

    def _read_xlsx(self, content: bytes) -> list[list[str]]:
        workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if cell is None else str(cell).strip() for cell in row])
        workbook.close()
        return [row for row in rows if any(cell for cell in row)]

    def _auto_map_columns(self, headers: list[str]) -> dict[str, str | None]:
        normalized_headers = {header: self._normalize_header(header) for header in headers}
        mapping: dict[str, str | None] = {}
        used_fields: set[str] = set()

        for header in headers:
            normalized = normalized_headers[header]
            matched_field: str | None = None
            for field, aliases in _FIELD_ALIASES.items():
                if field in used_fields:
                    continue
                if normalized in aliases or any(alias in normalized for alias in aliases):
                    matched_field = field
                    used_fields.add(field)
                    break
            mapping[header] = matched_field

        return mapping

    def _row_from_mapping(
        self,
        headers: list[str],
        raw_row: list[str],
        mapping: dict[str, str | None],
    ) -> ProductImportRow | None:
        values: dict[str, str] = {}
        for index, header in enumerate(headers):
            field = mapping.get(header)
            if field is None:
                continue
            cell = raw_row[index] if index < len(raw_row) else ""
            if cell:
                values[field] = cell

        name = values.get("name", "").strip()
        price = self._parse_price(values.get("price"))
        if not name or price is None:
            return None

        name, size = normalize_product_name_and_size(name, self._optional_str(values.get("size")))

        return ProductImportRow(
            name=name,
            price=price,
            category=self._optional_str(values.get("category")),
            manufacturer=self._optional_str(values.get("manufacturer")),
            size=size,
        )

    def _format_product_name(self, name: str, size: str | None) -> str:
        cleaned_name = name.strip()
        cleaned_size = (size or "").strip()
        if cleaned_size:
            return f"{cleaned_name} - {cleaned_size}"
        return cleaned_name

    def _parse_price(self, value: Any) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, (int, float, Decimal)):
            try:
                parsed = Decimal(str(value))
            except InvalidOperation:
                return None
            return parsed if parsed >= 0 else None

        text = str(value).strip()
        if not text:
            return None

        text = text.replace("\u00a0", "").replace(" ", "")
        if text.count(",") == 1 and text.count(".") == 0:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")

        text = re.sub(r"[^\d.]", "", text)
        if not text:
            return None

        try:
            parsed = Decimal(text)
        except InvalidOperation:
            return None
        return parsed if parsed >= 0 else None

    def _optional_str(self, value: Any) -> str | None:
        if value is None:
            return None
        cleaned = str(value).strip()
        return cleaned or None

    def _normalize_header(self, header: str) -> str:
        return re.sub(r"\s+", " ", header.strip().lower())
